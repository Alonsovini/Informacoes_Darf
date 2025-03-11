"""
Automação para extrair os dados das Darf e adicionar na planilha em excel
"""


import fitz  # PyMuPDF
import re
import os
from openpyxl import Workbook

def setup_folders_and_file():
    base_folder = r"C:\Info Darf"
    sub_folder = os.path.join(base_folder, "Darf")
    excel_file = os.path.join(base_folder, "Dados Darf.xlsx")

    # Cria as pastas se não existirem
    if not os.path.exists(base_folder):
        os.makedirs(base_folder)
        print(f"Criada a pasta: {base_folder}")
    if not os.path.exists(sub_folder):
        os.makedirs(sub_folder)

    # Cria o arquivo Excel se não existir
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "DARF Data"
        ws.append(["Período de Apuração", "CNPJ", "Código da Receita", "Data de Vencimento",
                    "Valor do Principal", "Valor Da Multa", "Valor dos Juros", "Valor Total"])
        wb.save(excel_file)
        print(f"Criado o arquivo Excel: {excel_file}")

    return sub_folder, excel_file

# Função para extrair informações de um PDF da DARF
def extract_darf_data(pdf_path):
    try:
        # Abre o PDF
        doc = fitz.open(pdf_path)
        text = ""

        # Extrai o texto do PDF (supondo que tenha apenas uma página)
        for page in doc:
            text += page.get_text()

        # Diagnóstico: mostra o texto extraído
        print(f"\nTexto extraído do arquivo {os.path.basename(pdf_path)}:\n{text}\n")

        # Expressões regulares ajustadas (ajuste conforme o texto extraído)
        data = {}

        # Primeiro formato
        principal_multa_juros_total = re.findall(
            r"AUTENTICAÇÃO BANCÁRIA.*?\n(\d{1,3}(?:\.\d{3})*,\d{2})\n(\d{1,3}(?:\.\d{3})*,\d{2})\n(\d{1,3}(?:\.\d{3})*,\d{2}).*?ATENÇÃO.*?(\d{1,3}(?:\.\d{3})*,\d{2})",
            text,
            re.DOTALL
        )

        if principal_multa_juros_total:
            data["Valor_Principal"] = principal_multa_juros_total[0][0]
            data["Valor_Multa"] = principal_multa_juros_total[0][1]
            data["Valor_Juros"] = principal_multa_juros_total[0][2]
            data["Valor_Total"] = principal_multa_juros_total[0][3]
        else:
            # Segundo formato
            valores = re.findall(
                r"ATENÇÃO\n([\d.,]+)\n([\d.,]+)\n([\d.,]+)\n([\d.,]+)\n([\d.,]+)",
                text
            )
            if valores:
                data["Valor_Principal"] = valores[0][0]
                data["Valor_Multa"] = valores[0][1]
                data["Valor_Juros"] = valores[0][2]
                data["Valor_Total"] = valores[0][3]
                data["Valor_Outro"] = valores[0][4]
            else:
                data["Valor_Principal"] = None
                data["Valor_Multa"] = None
                data["Valor_Juros"] = None
                data["Valor_Total"] = None
                data["Valor_Outro"] = None

        # Informações gerais (comuns aos dois formatos)
        data["Período de Apuração"] = re.findall(r"(\d{2}/\d{2}/\d{4})", text)[-1] if re.findall(r"(\d{2}/\d{2}/\d{4})", text) else None
        data["CNPJ"] = re.findall(r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", text)[-1] if re.findall(r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})", text) else None
        data["Codigo_Receita"] = re.findall(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\s*(\d{4})", text)[-1] if re.findall(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\s*(\d{4})", text) else None
        data["Data_Vencimento"] = re.findall(r"(\d{2}/\d{2}/\d{4})", text)[-2] if len(re.findall(r"(\d{2}/\d{2}/\d{4})", text)) > 1 else None

        doc.close()
        return data
    except Exception as e:
        print(f"Erro ao processar o PDF {pdf_path}: {e}")
        return None


# Função principal para processar todos os PDFs em uma pasta
def process_darf_pdfs(pdf_folder, output_excel):

    # Cria um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "DARF Data"
    ws.append(["Data de Vencimento", "CNPJ", "Código da Receita", "Período de Apuração",
               "Valor do Principal", "Valor Da Multa", "Valor dos Juros", "Valor Total"])

    # Processa cada PDF na pasta
    row_number = 2  # Começa da segunda linha (A primeira tem cabeçalho)
    for file in os.listdir(pdf_folder):
        if file.lower().endswith(".pdf"):
            pdf_path = os.path.join(pdf_folder, file)
            print(f"Processando: {pdf_path}")

            # Extrai os dados
            data = extract_darf_data(pdf_path)
            if data:
                # Adiciona os dados ao Excel
                ws.append([
                    data.get("Período de Apuração"),
                    data.get("CNPJ"),
                    data.get("Codigo_Receita"),
                    data.get("Data_Vencimento"),
                    data.get("Valor_Principal"),
                    data.get("Valor_Multa"),
                    data.get("Valor_Juros"),
                    None  # Placeholder para a fórmula
                ])

                ws[f"H{row_number}"] = f"=E{row_number}+G{row_number}+F{row_number}"
                row_number += 1

    # Salva o arquivo Excel
    wb.save(output_excel)
    print(f"Dados exportados para {output_excel}")

# Executa o script
pdf_folder, output_excel = setup_folders_and_file()
process_darf_pdfs(pdf_folder, output_excel)